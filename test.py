from openpyxl import Workbook, load_workbook

wb = Workbook()
wb = load_workbook('excel_test.xlsx')
sheet = wb.active

#
# print(sheet["A3"].value)
# print(sheet.cell(row=3, column=2).value)
#
# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=2):
#     for cell in row:
#         print(cell.value)
#
# sheet = wb.get_sheet_by_name("Arkusz1")
# print(sheet["C3"].value)
#
# print(wb.get_sheet_names())

# call = sheet['A7']
# call.value = "=SUM(A1:A6)"

# sheet["A3"] = "Witam"
# sheet.cell(row=3, column=2).value = 2022
# zbior = [1, 2, 3, 4, 45, 66, 777, 888]
# sheet.append(zbior)
#
# wb.save('excel_test.xlsx')
